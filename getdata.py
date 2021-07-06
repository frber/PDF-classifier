import os
import pdfplumber
import shutil
import nltk
from nltk.corpus import stopwords
import openpyxl
import re
from string import digits



def till_excel(text, fil_clean, ws):
    ws.cell(row=ws.max_row+1, column=1).value = text
    ws.cell(row=ws.max_row, column=2).value = fil_clean


def rensa_text(text, fil_clean, ws):
    text = ''.join([i for i in text if not i.isdigit()]) # Ta bort siffror
    stoppord = set(stopwords.words('swedish'))
    text = ' '.join(word for word in text.split() if word not in stoppord) # Ta bort stoppord
    till_excel(text, fil_clean, ws)


def hamta_text(filvag):
    try:
        open(filvag, 'r')
    except:
        return "fel"

    lista_text = []
    with pdfplumber.open(filvag) as pdf:
        sidor = pdf.pages
        for sida in sidor:
            text = sida.extract_text()
            lista_text.append(text)

    fulltext = ' '.join(lista_text)
    return fulltext


def sok_pdf(filvag, ws):
    for root, dirs, files in os.walk(filvag):
        for fil in files:
            if fil.endswith(".pdf"):
                filvag = os.path.join(root, fil)
                text = hamta_text(filvag)
                if text == "fel":
                    print("FEL", filvag)
                    continue
                else:
                    bas = os.path.basename(filvag)
                    fil = os.path.splitext(bas)[0]
                    ta_bort = str.maketrans('', '', digits)
                    fil_clean = fil.translate(ta_bort)
                    fil_clean = fil_clean.lower()
                    rensa_text(text, fil_clean, ws)


def main():
    orignal = r'Docs\Orginal\Data.xlsx'
    output = r'Docs\Data.xlsx'
    shutil.copy(orignal, output)
    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb['Data']
    filvag = r'pdftraindocs'
    sok_pdf(filvag, ws)
    wb.save(output)
    wb.close()
    os.startfile('Docs\Data.xlsx')


if __name__ == "__main__":
    main()


